# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
from django.utils.text import slugify
from datetime import datetime, date, timedelta
from django import template
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from django.shortcuts import render
from django.db import connection
from django.conf import settings
from django.db.models import Sum
import random
#importe letras
import tempfile
import os
import json
import openpyxl
from decimal import Decimal
from openpyxl.utils import get_column_letter


from .existing_models import Contratos, ContratoParcelas, Pessoas, Eventos
#from .forms import CAD_ClienteForm, Calculo_RepasseForm
from .models import Calculo_Repasse, CadCliente, Debito, Credito, Taxa, RepasseRetido, Dado

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, date):
            return obj.strftime('%Y-%m-%d')
        return super().default(obj)


@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}

    html_template = loader.get_template('home/index.html')
    return HttpResponse(html_template.render(context, request))


def preencher_tabela_cob(data_inicio, data_fim):
    with connection.cursor() as cursor:
        cursor.execute(f"""
select
dado.id_vendedor as codigo_vendedor,
dado.id_contrato,
dado.vendedor,
dado.id_comprador,
dado.comprador,
dado.parcelas_contrato,
dado.vl_pago,
dado.dt_vencimento,
dado.dt_credito,
dado.banco,
dado.contrato,
dado.evento,
dado.deposito,
dado.calculo,
dado.taxas,
dado.adi,
dado.me,
dado.op,
dado.repasses,
dado.comissao,
sum(dado.repasses) as total_repasse,
dado.id as dado_id
from dado
where dado.dt_credito between "2023-02-1" and "2023-02-03"
group by dado.id_contrato
""")
        result = cursor.fetchall()
        return result


@login_required(login_url="/login/")
def pages(request):
    context = {}

    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:
        request.session['serialized_data'] = None
        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        if load_template == 'tbl_comissoes_bootstrap.html':
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(vl_pago) as total_repasses, comissao, COUNT(*) as qtde
                    FROM calculo_repasse
                    WHERE comissao IS NOT NULL AND comissao != " -" AND comissao != "-" AND comissao != "            -" AND comissao != " - "
                    #AND dt_credito BETWEEN '2022-08-12' AND '2022-01-23'
                    group by comissao""")
                result = cursor.fetchall()
                context['sql'] = result

        elif load_template == 'tbl_bootstrap.html':
            if request.method == 'POST':
                pass
        

        elif load_template == 'cad_clientes_table_bootstrap.html':
            context['cad_clientes'] = CadCliente.objects.all()
            
        elif load_template == 'tbl_julia_bootstrap.html':
            if request.method == 'POST':
                bancos = request.POST.get('bancos')
                data = request.POST.get('data')
                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        select distinct comissao as comissionista,
                        sum(op) as comissoes
                        from calculo_repasse
                        where dt_credito = '{data}' and banco='{str(bancos).upper()}' and not isnull(comissao)
                        group by comissao;
                        """)
                    context['comissoes'] = cursor.fetchall()
                    cursor.execute(f"""
                        select id_contrato_id,
                        pessoas.nome,
                        CASE WHEN id_contrato_id > 12460 OR ISNULL(id_contrato_id)
                        THEN 'UNICRED' ELSE 'BRADESCO' END as banco,
                        sum(repasses)
                        from calculo_repasse
                        left join contratos on contratos.id=id_contrato_id
                        left join pessoas on pessoas.id = contratos.vendedor_id
                        where dt_credito = '{data}'
                        group by contratos.vendedor_id
                    """)
                    context['repasses'] = cursor.fetchall()
                    cursor.execute(f"""
                        select vendedor_id, pessoas.nome, total_repasses
                        from cad_cliente
                        left join pessoas on pessoas.id = vendedor_id
                        left join (
                            select id_vendedor, sum(coalesce(repasses,0)) as total_repasses
                            from dado
                            where dado.dt_credito = "{data}"
                            group by dado.id_vendedor
                        ) as dado on dado.id_vendedor = cad_cliente.vendedor_id
                        where repasse_semanal = true""")
                    context['repasses_semanais'] = cursor.fetchall()
                    
                context['valores_pagos_honorarios'] = Dado.objects.filter(dt_credito=data, 
                    banco='UNICRED').aggregate(
                        valores_pagos=Sum('vl_pago'),
                        honorarios=Sum('me')
                    )
                context['comissionistas_do_mes'] = Dado.objects.filter(
                    dt_credito=data,comissao__isnull=False
                    ).values('comissao').annotate(comissoes=Sum('op'))
                context['repasses_geral'] = Dado.objects.filter(dt_credito=data, banco=bancos).aggregate(
                    repasses=Sum('repasses')
                )
                context['repasses_semanais_vendedores_totais'] = sum([(querie[2] or 0) for querie in context['repasses_semanais']])
                context['repasses_gera_descontado'] = (context['repasses_geral']['repasses'] or 0) - (context['repasses_semanais_vendedores_totais'] or 0)
                    #context['repasses_geral'] = sum([float(calculo_repase.repasses) for calculo_repase in Calculo_Repasse.objects.filter(dt_credito=data, banco=bancos)])

            
        elif load_template == 'form_elements.html':
            if request.method == "POST":
                data_inicio = request.POST.get('data_inicio')  # 2022-08-01:str
                data_fim = request.POST.get('data_fim')  # 2022-08-21:str
                context['sql'] = preencher_tabela_cob(
                    data_inicio=data_inicio, data_fim=data_fim)
                request.session['serialized_data'] = json.dumps(context['sql'], cls=CustomJSONEncoder)
        
        elif load_template == 'tbl_credito.html':
            if request.method == 'POST':
                if 'novo-credito' in request.POST:
                    credor = request.POST.get('credor')
                    pagador = request.POST.get('pagador')
                    valor = request.POST.get('valor')
                    data_credito = request.POST.get('data-credito')
                    descricao = request.POST.get('descricao')
                    if credor == pagador:
                        return HttpResponse("Credor e Pagador não podem ser iguais")
                    if pagador:
                        Debito.objects.create(cliente=Pessoas.objects.get(id=pagador), vl_debito=valor, dt_debitado=data_credito, descricao=descricao)
                    Credito.objects.create(cliente=Pessoas.objects.get(id=credor), vl_credito=valor, dt_creditado=data_credito, descricao=descricao)
                elif 'filtrar-credito' in request.POST:
                    data_inicio = request.POST.get('data-inicio')  # 2022-08-01:str
                    data_fim = request.POST.get('data-fim')  # 2022-08-21:str
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    dias_de_consulta = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    dias = []
                    for dia in dias_de_consulta:
                        dias.append(f"SUM(CASE WHEN DAY(credito.dt_creditado) = {dia} THEN credito.vl_credito ELSE 0 END) AS dia_{dia}")
                    context['dias'] = dias_de_consulta
                    with connection.cursor() as cursor:
                        cursor.execute(f"""
                            SELECT cliente_id, nome,
                            {', '.join(dias)},
                            SUM(credito.vl_credito) as total_credito
                            from credito
                            left join pessoas on pessoas.id = cliente_id
                            where date(credito.dt_creditado) >= '{data_inicio}' AND date(credito.dt_creditado) <= '{data_fim}'
                            group by credito.cliente_id
                            order by cliente_id""")
                        context['creditos'] = cursor.fetchall()
                    tbody = ""
                    for resultado in context['creditos']:
                        cliente_id = resultado[0]
                        nome = resultado[1]
                        valores_diarios = resultado[2:-1]
                        total_credito = resultado[-1]
                        linha = f"""
                        <tr>
                            <td>{cliente_id}</td>
                            <td>{nome}</td>
                        """
                        for dia in valores_diarios:
                            linha += f"<td>{dia}</td>"
                        linha += f"<td>{total_credito}</td></tr>"
                        tbody += linha
                    context['tbody'] = tbody
                    
        elif load_template == 'tbl_debito.html':
            if request.method == 'POST':
                if 'novo-debito' in request.POST:
                    pagador = request.POST.get('pagador')
                    credor = request.POST.get('credor')
                    valor = request.POST.get('valor')
                    data_debito = request.POST.get('data-debito')
                    descricao = request.POST.get('descricao')
                    if credor == pagador:
                        return HttpResponse("Credor e Pagador não podem ser iguais")
                    if credor:
                        Credito.objects.create(cliente=Pessoas.objects.get(id=credor), vl_credito=valor, dt_creditado=data_debito, descricao=descricao)
                    Debito.objects.create(cliente=Pessoas.objects.get(id=pagador), vl_debito=valor, dt_debitado=data_debito, descricao=descricao)
                elif 'filtrar-debito' in request.POST:
                    data_inicio = request.POST.get('data-inicio')  # 2022-08-01:str
                    data_fim = request.POST.get('data-fim')  # 2022-08-21:str
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    dias_de_consulta = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    dias = []
                    for dia in dias_de_consulta:
                        dias.append(f"SUM(CASE WHEN DAY(debito.dt_debitado) = {dia} THEN debito.vl_debito ELSE 0 END) AS dia_{dia}")
                    context['dias'] = dias_de_consulta
                    with connection.cursor() as cursor:
                        cursor.execute(f"""
                            SELECT cliente_id, nome,
                            {', '.join(dias)},
                            SUM(debito.vl_debito) as total_debito
                            from debito
                            left join pessoas on pessoas.id = debito.cliente_id
                            where date(debito.dt_debitado) >= '{data_inicio}' AND date(debito.dt_debitado) <= '{data_fim}'
                            group by debito.cliente_id
                            order by cliente_id""")
                        context['debitos'] = cursor.fetchall()
                    tbody = ""
                    for resultado in context['debitos']:
                        cliente_id = resultado[0]
                        nome = resultado[1]
                        valores_diarios = resultado[2:-1]
                        total_debito = resultado[-1]
                        linha = f"""
                        <tr>
                            <td>{cliente_id}</td>
                            <td>{nome}</td>
                        """
                        for dia in valores_diarios:
                            linha += f"<td>{dia}</td>"
                        linha += f"<td>{total_debito}</td></tr>"
                        tbody += linha
                    context['tbody'] = tbody
                    
        elif load_template == 'tbl_taxas.html':
            if request.method == 'POST':
                if 'nova-taxa' in request.POST:
                    cliente_id = request.POST.get('cliente')
                    valor = request.POST.get('valor')
                    data_taxa = request.POST.get('data-taxa')
                    descricao = request.POST.get('descricao')
                    try:
                        cliente = Pessoas.objects.get(id=cliente_id)
                    except Pessoas.DoesNotExist:
                        return HttpResponse("Cliente não encontrado")
                    except Pessoas.MultipleObjectsReturned:
                        return HttpResponse("Mais de um cliente encontrado, {}".format(Pessoas.objects.filter(id=cliente_id)))
                    Taxa.objects.create(cliente=cliente, taxas=valor, dt_taxa=data_taxa, descricao=descricao)
                elif 'filtrar-taxa' in request.POST:
                    data_inicio = request.POST.get('data-inicio')  # 2022-08-01:str
                    data_fim = request.POST.get('data-fim')  # 2022-08-21:str
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    dias_de_consulta = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    dias = []
                    for dia in dias_de_consulta:
                        dias.append(f"SUM(CASE WHEN DAY(taxas.dt_taxa) = {dia} THEN taxas.taxas ELSE 0 END) AS dia_{dia}")
                    context['dias'] = dias_de_consulta
                    with connection.cursor() as cursor:
                        cursor.execute(f"""
                            SELECT cliente_id, nome,
                            {', '.join(dias)},
                            SUM(taxas.taxas) as total_taxas
                            from taxas
                            left join pessoas on pessoas.id = taxas.cliente_id
                            where date(taxas.dt_taxa) >= '{data_inicio}' AND date(taxas.dt_taxa) <= '{data_fim}'
                            group by taxas.cliente_id
                            order by cliente_id""")
                        context['taxas'] = cursor.fetchall()
                    tbody = ""
                    for resultado in context['taxas']:
                        cliente_id = resultado[0]
                        nome = resultado[1]
                        valores_diarios = resultado[2:-1]
                        total_taxas = resultado[-1]
                        linha = f"""
                        <tr>
                            <td>{cliente_id}</td>
                            <td>{nome}</td>
                        """
                        for dia in valores_diarios:
                            linha += f"<td>{dia}</td>"
                        linha += f"<td>{total_taxas}</td></tr>"
                        tbody += linha
                    context['tbody'] = tbody
                    
        elif load_template == 'tbl_repasse_retido.html':
            if request.method == 'POST':
                if 'novo-repasse-retido' in request.POST:
                    cliente_id = request.POST.get('cliente')
                    valor = request.POST.get('valor')
                    data_repasse_retido = request.POST.get('data-repasse-retido')
                    tipo = request.POST.get('tipo')
                    try:
                        cliente = Pessoas.objects.get(id=cliente_id)
                        RepasseRetido.objects.create(cliente=cliente, vlr_rep_retido=valor, dt_rep_retido=data_repasse_retido, tipo=tipo)
                    except Pessoas.DoesNotExist:
                        return HttpResponse("Cliente não encontrado")
                    except Pessoas.MultipleObjectsReturned:
                        return HttpResponse("Mais de um cliente encontrado, {}".format(Pessoas.objects.filter(id=cliente_id)))
                elif 'filtrar-repasse-retido' in request.POST:
                    data_inicio = request.POST.get('data-inicio')  # 2022-08-01:str
                    data_fim = request.POST.get('data-fim')  # 2022-08-21:str
                    data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
                    data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
                    dias_de_consulta = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]
                    dias = []
                    for dia in dias_de_consulta:
                        dias.append(f"SUM(CASE WHEN DAY(repasse_retido.dt_rep_retido) = {dia} THEN repasse_retido.vlr_rep_retido ELSE 0 END) AS dia_{dia}")
                    context['dias'] = dias_de_consulta
                    with connection.cursor() as cursor:
                        cursor.execute(f"""
                            SELECT cliente_id, nome,
                            {', '.join(dias)},
                            SUM(repasse_retido.vlr_rep_retido) as total_repasse_retido
                            from repasse_retido
                            left join pessoas on pessoas.id = repasse_retido.cliente_id
                            where date(repasse_retido.dt_rep_retido) >= '{data_inicio}' AND date(repasse_retido.dt_rep_retido) <= '{data_fim}'
                            group by repasse_retido.cliente_id
                            order by cliente_id""")
                        context['repasse_retido'] = cursor.fetchall()
                    tbody = ""
                    for resultado in context['repasse_retido']:
                        cliente_id = resultado[0]
                        nome = resultado[1]
                        valores_diarios = resultado[2:-1]
                        total_repasse_retido = resultado[-1]
                        linha = f"""
                        <tr>
                            <td>{cliente_id}</td>
                            <td>{nome}</td>
                        """
                        for dia in valores_diarios:
                            linha += f"<td>{dia}</td>"
                        linha += f"<td>{total_repasse_retido}</td></tr>"
                        tbody += linha
                    context['tbody'] = tbody
                
        elif load_template == 'tbl_debito_cessao.html':
            if request.method == 'POST':
                pass
                context['nothing'] = None
                
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT distinct vendedor_id, nome , nu_parcelas, dt_contrato FROM contratos
                    join pessoas
                    where (dt_contrato >= '2022-09-01' and dt_contrato <= '2022-09-30')
                    and repasse = 'S' and pessoas.id=vendedor_id
                    order by contratos.id desc;
                    """
                )
                context['sql'] = cursor.fetchall()
                
        elif load_template == 'pessoa_info.html':
            if request.method == 'POST':
                pass
            context['pessoa'] = Pessoas.objects.get(id=733 or None)
                
            

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    """ except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request)) """


def consulta_por_data(request):
    context: dict = {}
    if request.method == 'POST':
        data_inicio = request.POST.get('data-inicio')  # 2022-08-01:str
        data_fim = request.POST.get('data-fim')  # 2022-08-21:str
        date_data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        date_date_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        #selecione todos os contratos em que as suas parcelas estejam dentro do intervalo de datas
        #contratos = Contratos.objects.filter(parcelas__dt_credito__range=(date_data_inicio, date_date_fim))[:30]
        contratos = Contratos.objects.filter(
            parcelas__dt_credito__gte=date_data_inicio, parcelas__dt_credito__lte=date_date_fim)[:100]
        parcelas = ContratoParcelas.objects.filter(
            dt_credito__gte=date_data_inicio, dt_credito__lte=date_date_fim)[:100]
        # Cannot filter a query once a slice has been taken.
        #parcelas = contratos.filter(dt_credito__gte=date_data_inicio, dt_credito__lte=date_date_fim)

        context = {'contratos': contratos, 'parcelas': parcelas}

        return render(request, 'home/tbl_bootstrap.html', context=context)
    return HttpResponse('<h1>GET</h1>')


def criar_cad_cliente(request):
    context = {}
    """ Aplicar os calculos da planilha aqui """
    if request.method == "POST":
        return HttpResponse('<h1>POST</h1>')
    with connection.cursor() as cursor:
        cursor.execute("SELECT tp_contrato, status ,vl_boleto, vl_pago, vl_parcela, nu_parcelas, dados_arquivo_retorno.dt_credito, contrato_parcelas.dt_credito FROM dados_arquivo_retorno, contratos, contrato_parcelas LIMIT 10")
        result = cursor.fetchall()
        context['sql'] = result
    return HttpResponse('<h1>GET, {}</h1>'.format(context['sql']))

def criar_novo_cadastro_de_credito_e_debito(request, *args, **kwargs):
    #pegue os valores vindo do form, eles são: credor, pagador, valor, data-credito e descricao
    if request.method == 'POST':
        credor = request.POST.get('credor')
        pagador = request.POST.get('pagador')
        valor = request.POST.get('valor')
        data_credito = request.POST.get('data-credito')
        #data_credito = datetime.strptime(request.POST.get('data-credito'), '%Y-%m-%d').date()
        descricao = request.POST.get('descricao')
        pagador = Pessoas.objects.get(id=pagador)
        credor = Pessoas.objects.get(id=credor)
        """ Debito.objects.create(
            cliente = pagador,
            vl_debito = valor,
            dt_debitado = data_credito,
            descricao = descricao,
        ) """
        """ Credito.objects.create(
            cliente = credor,
            vl_credito = valor,
            dt_creditado = data_credito,
            descricao = descricao,
        ) """
        return HttpResponseRedirect('/tbl_credito_cessao.html')
    return HttpResponse("<h1>GET OR ANY REQUEST</h1>")

def criar_nova_taxa(request):
    if request.method == 'POST':
        try:
            cliente = Pessoas.objects.get(id=request.POST.get('id-cliente'))
        except Pessoas.DoesNotExist:
            return HttpResponse("<h1>Pessoa não Encontrada</h1>")
        except Pessoas.MultipleObjectsReturned:
            return HttpResponse("<h1>Erro: Mais de uma pessoa encontrada</h1>")
        taxas = request.POST.get('taxas')
        tipo = request.POST.get('tipo')
        descricao_taxa = request.POST.get('descricao-taxa')
        data_taxa = request.POST.get('data-taxa')
        """ Taxa.objects.create(
            cliente = cliente,
            tipo = tipo,
            descricao = descricao_taxa,
            taxas = taxas,
            dt_taxa = data_taxa
        ) """
        return HttpResponseRedirect('/tbl_credito_cessao.html')
    return HttpResponse("<h1>GET OR ANY REQUEST</h1>")

def criar_novo_repasse_retido(request, *args, **kwargs):
    if request.method == 'POST':
        id_cliente = request.POST.get('id-cliente')
        valor = request.POST.get('valor')
        tipo = request.POST.get('tipo')
        data_repasse_retido = request.POST.get('data-repasse-retido')
        """ RepasseRetido.objects.create(
            cliente=Pessoas.objects.get(id=id_cliente),
            vlr_rep_retido=valor,
            tipo = tipo,
            dt_rep_retido = data_repasse_retido
        ) """
        return HttpResponseRedirect('/tbl_credito_cessao.html')
    return HttpResponse(" <h1>GET</h1> ")

def filtrar_tabela_quinzenal(request, *args, **kwargs):
    context:dict = dict()
    if request.method == 'POST':
        data_inicio = request.POST.get('data-inicio')
        data_fim = request.POST.get('data-fim')
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
        
        dias_de_consulta = [(data_inicio_dt + timedelta(days=x)).day for x in range((data_fim_dt - data_inicio_dt).days + 1)]

        dias = []
        for dia in dias_de_consulta:
            dias.append(f"SUM(CASE WHEN DAY(dado.dt_credito) = {dia} THEN dado.repasses ELSE 0 END) AS dia_{dia}")
            
        consulta = f"""
        select
dado.id_vendedor,
dado.vendedor,
coalesce(total_repasse_retido,0) as total_repasse_retido,
{', '.join(dias)},
coalesce(total_credito,0) as total_credito,
coalesce(total_debito,0) as total_debito,
coalesce(total_taxas,0) as total_taxas,
sum(dado.repasses)
	+ coalesce(total_repasse_retido, 0)
    + coalesce(total_credito, 0)
    - coalesce(total_debito, 0)
    - coalesce(total_taxas, 0)
    as total_repasse
from dado
left join (
	select sum(vl_credito) as total_credito, cliente_id
    from credito
    where credito.dt_creditado between "{data_inicio}" and "{data_fim}"
    group by cliente_id
) as credito on credito.cliente_id = dado.id_vendedor
left join (
	select sum(vl_debito) as total_debito, cliente_id
    from debito
    where debito.dt_debitado between "{data_inicio}" and "{data_fim}"
    group by cliente_id
) as debito on debito.cliente_id = dado.id_vendedor
left join (
	select sum(taxas) as total_taxas, cliente_id
    from taxas
    where taxas.dt_taxa between "{data_inicio}" and "{data_fim}"
    group by cliente_id
) as taxas on taxas.cliente_id = dado.id_vendedor
left join (
	select sum(vlr_rep_retido) as total_repasse_retido, cliente_id
    from repasse_retido
    where repasse_retido.dt_rep_retido between "{data_inicio}" and "{data_fim}"
    group by cliente_id
) as repasse_retido on repasse_retido.cliente_id = dado.id_vendedor
where date(dado.dt_credito) BETWEEN '{data_inicio}' AND '{data_fim}'
group by dado.id_vendedor
order by dado.id_vendedor
        """
    
        with connection.cursor() as cursor:
            cursor.execute(consulta)
            context['data'] = cursor.fetchall()
        
        request.session['serialized_data'] = json.dumps(context['data'], cls=CustomJSONEncoder)
        
        context['dias_de_consulta'] = dias_de_consulta
        tbody = ""
        for resultado in context['data']:
            vendedor_id = resultado[0]
            nome_vendedor = resultado[1]
            valor_repasse_retido = resultado[2]
            #dt_credito = resultado[3]
            valores_diarios = resultado[3:-4]
            total_creditos = resultado[-4]
            total_taxas = resultado[-3]
            totaL_debitos = resultado[-2]
            total_repasses = resultado[-1]

            linha = f"""
            <tr>
                <td>{vendedor_id}</td>
                <td>{nome_vendedor}</td>
                <td>{valor_repasse_retido}</td>
            """
            for valor_dia in valores_diarios:
                linha += f"<td>{valor_dia}</td>"
            
            linha += f"""
            <td>{total_creditos}</td>
            <td>{totaL_debitos}</td>
            <td>{total_taxas}</td>
            <td>{total_repasses}</td>
            </tr>
            """
            tbody += linha

        context['tbody'] = tbody
        return render(request, 'home/tbl_bootstrap.html', context=context)
    return HttpResponse(" <h1>GET OR ANY REQUEST</h1> ")

def upload_planilha_quinzenal(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        #arquivo esta recebendo com sucesso, azer os devidos tratamentos
        return HttpResponse(planilha)
    return HttpResponseRedirect('/tbl_credito_cessao.html')

def download_planilha_quinzenal(request, *args, **kwargs):
    if request.session.get('serialized_data') is None:
        return HttpResponse('Nenhum dado encontrado para download')
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, f'planilha_consulta_quinzenal_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'ID Vendedor'
        sheet['B1'] = 'Nome Vendedor'
        sheet['C1'] = 'Valor Repasse Retido'
        
        # define a largura das primeiras colunas
        sheet.column_dimensions[get_column_letter(1)].width = 15
        sheet.column_dimensions[get_column_letter(2)].width = 30
        sheet.column_dimensions[get_column_letter(3)].width = 30
        
        for i, row in enumerate(json.loads(request.session.get('serialized_data'))):
            for j, value in enumerate(row):
                sheet.cell(row=i+2, column=j+1, value=value)
        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response

def download_planilha_cob(request, *args, **kwargs):
    if request.session.get('serialized_data') is None:
        return HttpResponse('Nenhum dado encontrado para download')
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, f'planilha_consulta_cob_{slugify(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        """ 
        <th>Vendedor ID</th>
        <th>Contrato ID</th>
        <th>Vendedor</th>
        <th>Comprador ID</th>
        <th>Comprador</th>
        <th>Parcelas Contrato</th>
        <th>Valor Pago</th>
        <th>Data Vencimento</th>
        <th>Data Credito</th>
        <th>Banco</th>
        <th>Contrato</th>
        <th>Evento</th>
        <th>Deposito</th>
        <th>Calculo</th>
        <th>Taxas</th>
        <th>ADI</th>
        <th>ME</th>
        <th>OP</th>
        <th>Repasses</th>
        <th>Comissao</th>
        <th>Total Repasse</th>
        <th>Dado ID</th>
        """
        lista_header_consulta_sql = ['codigo_vendedor', 'id_contrato', 'vendedor', 'id_comprador', 'comprador', 'parcelas_contrato', 'vl_pago', 'dt_vencimento', 'dt_credito', 'banco', 'contrato', 'evento', 'deposito', 'calculo', 'taxas', 'adi', 'me', 'op', 'repasses', 'comissao', 'total_repasse', 'dado_id']
        for i, value in enumerate(lista_header_consulta_sql):
            sheet[get_column_letter(i+1)+'1'] = str(value)
        
        
        for i, row, in enumerate(json.loads(request.session.get('serialized_data'))):
            for j, value in enumerate(row):
                sheet.cell(row=i+2, column=j+1, value=value)
        workbook.save(filepath)
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
            return response

def upload_planilha_cob(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        #verificar se o arquivo é do tipo xlsx
        if planilha.name.endswith('.xlsx'):
            #arquivo esta recebendo com sucesso, azer os devidos tratamentos
            wb = openpyxl.load_workbook(planilha)
            #nesse arquivo de planilha, a primeira aba é a que contem os dados, e so existe uma aba
            sheet = wb.worksheets[0]
            linha = 0
            for row in sheet.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                linha += 1
                if (row[0] and row[1] and row[2]) == None:
                    break
            
            
        #arquivo esta recebendo com sucesso, azer os devidos tratamentos
        return HttpResponse(planilha)
    return HttpResponseRedirect('/form_elements.html')

def editar_boleto_avulso(request, *args, **kwargs):
    return HttpResponse("<h1>FUNCIONOU</h1>")

def upload_planilha_cavalos_cob(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif planilha.name.endswith('.xlsx'):
            pessoa_nula = Pessoas.objects.get(id=99999)
            erros:list[str] = []
            linhas_nulas = 0
            wb = openpyxl.load_workbook(planilha)
            cob = wb.active
            linha = 0
            for row in cob.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                if (row[0] and row[1] and row[2]) == (None or ''):
                    linhas_nulas += 1
                    if linhas_nulas == 2:
                        break
                    continue
                codigo_vendedor = row[0]
                contrato_id = row[1]
                nome_vendedor = row[2]
                nome_comprador = row[3]
                parcela_paga = row[4]
                contrato_parcelas = row[5]
                valor_parcela = row[6]
                dt_vencimento = row[7]#dia/mes/ando
                """ dt_vencimento_dt = datetime.combine(date(day=dt_vencimento.day, 
                    month=dt_vencimento.month, 
                    year=dt_vencimento.year), datetime.min.time()) """
                dt_credito = row[8]
                """ dt_credito_dt = datetime.combine(date(day=dt_credito.day,
                    month=dt_credito.month,
                    year=dt_credito.year), datetime.min.time()) """
                banco = row[9]
                produto_ou_contrato = row[10]
                evento = row[11]
                deposito = row[12]
                calc = row[13]
                taxas = row[14]
                adi = row[15]
                me = row[16]
                op = row[17]
                repasses = row[18]
                comissao = row[19]
                try:
                    vendedor = Pessoas.objects.get(id=codigo_vendedor)
                except Pessoas.DoesNotExist:
                    vendedor = Pessoas.objects.create(id=codigo_vendedor, nome=nome_vendedor, email=f"{row[0]}_{row[2]}@gmail.com")
                except Pessoas.MultipleObjectsReturned:
                    erros.append(f"O vendedor {nome_vendedor} possui mais de um cadastro. linha: {linha}")
                    continue
                try:
                    comprador = Pessoas.objects.get(nome=nome_comprador)
                except Pessoas.DoesNotExist:
                    comprador = Pessoas.objects.create(nome=nome_comprador, email=f"{nome_comprador}@gmail.com")
                except Pessoas.MultipleObjectsReturned or Exception as e:
                    erros.append(f"O comprador {nome_comprador} possui mais de um cadastro. linha: {linha}")
                    continue
                    
                if row[1] is not None and type(row[1]) == int:
                    if row[1] == 1:
                        eventos = Eventos.objects.create(nome=row[4], leiloeiro=pessoa_nula, dt_evento=date.today(), tipo="qualquer")
                        Contratos.objects.create(vendedor=vendedor, comprador=comprador, eventos=eventos, descricao=row[10], pessoas_id_inclusao=pessoa_nula)
                    else:
                        try:
                            contratos = Contratos.objects.get(id=row[1])
                            eventos = Eventos.objects.get(id=contratos.eventos.id)
                            eventos.nome = evento
                            contratos.eventos = eventos
                            contratos.descricao = produto_ou_contrato
                            eventos.save()
                            contratos.save()
                        except Contratos.DoesNotExist:
                            eventos = Eventos.objects.create(nome=evento, leiloeiro=Pessoas.objects.get(id=99999),
                                dt_evento=date.today(), tipo="qualquer")
                            contratos = Contratos.objects.create(id=contrato_id,
                                vendedor=vendedor, comprador=comprador,
                                descricao=row[10], eventos=eventos, pessoas_id_inclusao=pessoa_nula)
                else:
                    erros.append(f"O contrato {row[1]} não existe para o comprador: {row[3]} e vendedor: {row[2]}. linha: {linha}")
                    continue
                try:
                    contrato_parcelas = ContratoParcelas.objects.get(contratos=contratos, 
                        nu_parcela=int(str(row[5].split('/')[0][1:]))
                        )
                    contrato_parcelas.dt_vencimento = row[7]
                    contrato_parcelas.vl_parcela = row[6]
                    contrato_parcelas.dt_credito = row[8]
                    contrato_parcelas.save()
                except ContratoParcelas.DoesNotExist:
                    contrato_parcelas = ContratoParcelas.objects.create(
                        contratos=contratos,
                        dt_vencimento=row[7], vl_parcela=row[6], dt_credito=row[8],
                        nu_parcela=int(str(row[5].split('/')[0][1:]))
                    )
                except ContratoParcelas.MultipleObjectsReturned:
                    erros.append(f"O contrato {contratos.id} possui mais de uma parcela com o numero {row[5]}. linha: {linha}")
                    #!chances muito poucas de acontecer, mas se acontecer, o sistema vai pegar o primeiro objeto
                    pass
                try:
                    calculo_repasse = Calculo_Repasse.objects.get(contrato_parcelas=contrato_parcelas)
                    calculo_repasse.deposito = row[12]
                    calculo_repasse.calculo = row[13]
                    calculo_repasse.taxas = row[14]
                    calculo_repasse.adi = row[15]
                    calculo_repasse.me = row[16]
                    calculo_repasse.op = row[17]
                    calculo_repasse.repasses = row[18]
                    calculo_repasse.comissao = row[19]
                    calculo_repasse.dt_credito = contrato_parcelas.dt_credito
                    calculo_repasse.save()
                except Calculo_Repasse.DoesNotExist:
                    calculo_repasse = Calculo_Repasse.objects.create(contrato_parcelas=contrato_parcelas,
                        deposito=row[12], calculo=row[13], taxas=row[14], adi=row[15],
                        me=row[16], op=row[17], repasses=row[18], comissao=row[19], dt_credito=contrato_parcelas.dt_credito
                    )
                    
                linha += 1
            return HttpResponse('Planilha de cavalos recebida com sucesso, linhas lidas: {}, erros: {}'.format(linha,erros))
        else:
            return HttpResponse('Arquivo não é do tipo xlsx')

def upload_planilha_cad_clientes(request, *args, **kwargs):
    if request.method == 'POST':
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif planilha.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(planilha)
            cad_cliente = wb.active
            linha = 0
            for row in cad_cliente.iter_rows(values_only=True):
                if linha < 1:
                    linha += 1
                    continue
                #* sistema de parada
                if (row[0]) == None:
                    break
                try:
                    pessoa = Pessoas.objects.get(id=row[2])
                    pessoa.nome = row[0]
                    pessoa.email = row[1]
                except Pessoas.DoesNotExist:
                    pessoa = Pessoas.objects.create(id=row[2], nome=row[0], email=row[1])
                except Pessoas.MultipleObjectsReturned:
                    pass
                pessoa.save()
                try:
                    cad_cliente = CadCliente.objects.get(vendedor=pessoa)
                    cad_cliente.sim = row[3]
                    cad_cliente.nao = row[4]
                    cad_cliente.operacional = row[5]
                    cad_cliente.tcc = row[6]
                    cad_cliente.honorarios = row[7]
                    cad_cliente.animal = row[8]
                    cad_cliente.evento = row[9]
                    cad_cliente.informar_repasse = row[10]
                except CadCliente.DoesNotExist:
                    cad_cliente = CadCliente.objects.create(vendedor=pessoa, sim=row[3], nao=row[4],
                        operacional=row[5], tcc=row[6], honorarios=row[7], animal=row[8],
                        evento=row[9], informar_repasse=row[10])
                except CadCliente.MultipleObjectsReturned:
                    pass
                cad_cliente.save()
                
            return HttpResponse('Planilha {} recebida com sucesso'.format(planilha.name))
        else:
            return HttpResponse('Arquivo não é do tipo xlsx')

def upload_planilha_parcelas_taxas(request, *args, **kwargs):
    if request.method == "POST":
        planilha = request.FILES.get('docpicker')
        if planilha is None:
            return HttpResponse('Nenhum arquivo selecionado')
        elif not planilha.name.endswith('.xlsx'):
            return HttpResponse('Arquivo não é do tipo xlsx')
        linhas_nulas = 0
        linhas = 0
        erros:list[str] = []
        wb = openpyxl.load_workbook(planilha)
        segunda_quinzena = wb.active
        for row in segunda_quinzena.iter_rows(values_only=True):
            if linhas < 1:
                linhas += 1
                continue
            if row[0] == None or row[0] == '':
                break
            contrato_parcelas_id = row[0]
            comprador_nome = row[1]
            vendedor_nome = row[2]
            parcela = row[3]# (1/2), pegar o primeiro digito antes da barra
            vencimento = row[4] #14/01/2023, dia 14 mes 01 ano 2023
            dt_vencimento = date(day=vencimento.day, month=vencimento.month, year=vencimento.year)
            dt_vencimento = datetime.combine(dt_vencimento, datetime.min.time())

            valor = row[5]
            tcc = row[6]
            ted = row[7]
            desconto_total = row[8]
            hon = row[9]
            repasse = row[10]
            try:
                try:
                    comprador = Pessoas.objects.get(nome=comprador_nome)
                except Pessoas.DoesNotExist:
                    comprador = Pessoas.objects.create(nome=comprador_nome, email=f"{comprador_nome}-email@nãoexiste.com.br")
                except Pessoas.MultipleObjectsReturned:
                    erros.append(f"Erro na linha {linhas}, comprador {comprador_nome} possui mais de um cadastro")
                Debito.objects.create(cliente=comprador, vl_debito=desconto_total, dt_debitado=dt_vencimento)
                try:
                    vendedor = Pessoas.objects.get(nome=vendedor_nome)
                except Pessoas.DoesNotExist:
                    vendedor = Pessoas.objects.create(nome=vendedor_nome, email=f"{vendedor_nome}-email@nãoexiste.com.br")
                except Pessoas.MultipleObjectsReturned:
                    erros.append(f"Erro na linha {linhas}, vendedor {vendedor_nome} possui mais de um cadastro")
                Credito.objects.create(cliente=vendedor, vl_credito=repasse, dt_creditado=dt_vencimento)
            except Exception as e:
                erros.append(f"Erro na linha {linhas}, {e}")
            linhas += 1
        
        return HttpResponse("Planilha Recebida com sucesso, linhas lidas, {}, erros: {}".format(linhas, erros))

def upload_planilha_dados_brutos(request):
    if request.method == "POST":
        if request.FILES.get('docpicker') is None:
            return HttpResponse("Nenhum arquivo selecionado")
        if not request.FILES.get('docpicker').name.endswith('.xlsx'):
            return HttpResponse("Arquivo não é do tipo xlsx")
        planilha = request.FILES.get('docpicker')
        wb = openpyxl.load_workbook(planilha)
        cob = wb.active
        linhas_nulas = 0
        linhas = 0
        erros:list[str] = []
        for row in cob.iter_rows(values_only=True):
            if linhas < 1:
                linhas += 1
                continue
            if (row[0] == None or row[0] == '') and (row[1] == None or row[1] == ''):
                linhas_nulas += 1
                if linhas_nulas > 1:
                    break
                continue
            linhas += 1
            vendedor_id = row[0]
            contrato_id = row[1]
            nome_vendedor = row[2]
            comprador = row[3]
            parcela_paga = row[4]
            parcelas_contrato = row[5]
            valor = row[6]
            data_vencimento = row[7]
            data_credito = row[8]
            banco = row[9]
            contrato = row[10]
            evento = row[11]
            deposito = row[12]
            calc = row[13]
            taxas = row[14]
            adi = row[15]
            me = row[16]
            op = row[17]
            repasses = row[18]
            comissao = row[19]
            id_excel = row[21]
            
            try:
                dado = Dado.objects.get(id_vendedor=vendedor_id, id_contrato=contrato_id, comprador=comprador, vl_pago=valor)
                """ altere todos os dados para a atual linha, exceto os id_vendedor=vendedor_id, id_contrato=contrato_id, comprador=comprador, vl_pago=valor """
                dado.nu_parcela = parcela_paga
                dado.parcelas_contrato = parcelas_contrato
                dado.dt_vencimento = data_vencimento
                dado.dt_credito = data_credito
                dado.banco = banco
                dado.contrato = contrato
                dado.evento = evento
                dado.deposito = deposito
                dado.calculo = calc
                dado.taxas = taxas
                dado.adi = adi
                dado.me = me
                dado.op = op
                dado.repasses = repasses
                dado.comissao = comissao
            except Dado.DoesNotExist:
                Dado.objects.create(
                id_vendedor=vendedor_id,
                id_contrato=contrato_id,
                vendedor=nome_vendedor,
                comprador=comprador,
                nu_parcela=parcela_paga,
                parcelas_contrato=parcelas_contrato,#(1/2) :str, pegar somente o primeiro digte antes da barra
                vl_pago=valor,
                dt_vencimento=data_vencimento,
                dt_credito=data_credito,
                banco=banco,
                contrato=contrato,
                evento=evento,
                deposito=deposito,
                calculo=calc,
                taxas=taxas,
                adi=adi,
                me=me,
                op=op,
                repasses=repasses,
                comissao=comissao
            )
            except Dado.MultipleObjectsReturned:
                erros.append(f"Foi encontrado mais de um dado com os mesmos dados na linha {linhas}\n chaves primarias de busca são: id_vendedor={vendedor_id}, id_contrato={contrato_id}, comprador={comprador}, vl_pago={valor}")
            except Exception as e:
                erros.append(f"Erro na linha {linhas}, Exception Error:{e}")
            
            
        erros_html = "<ul>"
        for erro in erros:
            erros_html += f"<li>{erro}</li>"
        erros_html = "</ul>"
        return HttpResponse("Planilha recebida com sucesso <br> linhas lidas: {} <br> erros: {}".format(linhas, erros_html))
    return HttpResponse("HTTP REQUEST")
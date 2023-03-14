import openpyxl
wb = openpyxl.load_workbook('cavalos_09_2022_backup.xlsx')
cob = wb['COB (2)']
def carregar_modeL_calcular_repasse():
   linha = 0
   for row in cob_2.iter_rows(values_only=True):
      if linha < 2:
            linha += 1
            continue
      if linha == 9908:
            break
      linha += 1
      try :
            Calculo_Repasse.objects.create(
            id_contrato=Contratos.objects.get(id=row[1]),
            vl_pago=row[6],
            deposito=row[12],
            taxas=row[14],
            adi=row[15],
            me=row[16],
            op=row[17],
            repasses=row[18],
            calculo=row[13],
            comissao=row[19]
         )
      except Exception as e:
         print(f'linha: {linha} - {e}')


def carregar_model_dados():
   linha = 0
   for row in cob_2.iter_rows(values_only=True):
      if linha < 2:
            linha += 1
            continue
      if linha == 9908:
            break
      linha += 1
      try:
         Dado.objects.create(
            id_vendedor=row[0],
            id_contrato=row[1],
            vendedor=row[2],
            comprador=row[3],
            nu_parcela=row[4],
            contrato=int(str(row[5].split('/')[0][1:])),#(19/30), pegar somente o numero do contrato, ou seja o 19
            vl_pago=row[6],
            dt_vencimento=row[7],
            dt_credito=row[8],
            banco=row[9],
            evento=row[10],
            deposito=row[12],
            calculo=row[13],
            taxas=row[14],
            adi=row[15],
            me=row[16],
            op=row[17],
            repasses=row[18],
            comissao=row[19]
         )
      except Exception as e:
         print(e)
         continue

def criar_calculos_repasse_com_base_nos_dados():
   for dado in Dado.objects.all():
      try:
         Calculo_Repasse.objects.create(
            id_contrato=Contratos.objects.get(id=dado.id_contrato),
            nu_parcela=dado.contrato,
            vl_pago=dado.vl_pago,
            deposito=dado.deposito,
            dt_credito=dado.dt_credito,
            taxas=dado.taxas,
            adi=dado.adi,
            me=dado.me,
            op=dado.op,
            repasses=dado.repasses,
            calculo=dado.calculo,
            comissao=dado.comissao
         )
      except Exception as e:
         print(e)
         continue

""" 
linhas_nulas = 0
                    linha = 0
                    for row in cad_cliente.iter_rows(values_only=True):
                        if linha < 1:#parametros para pular a primeira linha
                            linha += 1
                            continue
                        #* sistema de parada pronto!
                        if (row[0]) == None:
                            linhas_nulas += 1
                            if linhas_nulas == 2:
                                break
                            continue
                        return HttpResponse(row[0],row[1], row[2])
                        pessoas, pessoa_criada_bolean = Pessoas.objects.get_or_create(id=row[1], nome=row[0])
                        cad_cliente_obj, cad_cliente_criado_boolean = CadCliente.objects.get_or_create(
                            vendedor=pessoas, sim=row[2], nao=row[3], 
                            operacional=row[4], tcc=row[5], honorarios = row[6],
                            animal=row[7], evento=row[8], informar_repasse=row[9],
                        )
                        linha += 1
"""